# operaciones.py

import os
import pandas as pd
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox, simpledialog, filedialog
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from natsort import natsort_keygen

# ----------- DEFINIR EL MAPEO DE COLUMNAS GLOBALMENTE -----------
# Definir el mapeo de nombres de columnas
COLUMN_MAPPING = {
    'Bruto': 'Peso bruto (kg)',
    'Neto': 'Peso neto (kg)',
    'Volumen': 'Volumen (m3)',
    'Importe': 'Valor FOB',
    'Doc.comer.': 'Pedido',
    'LibrUtiliz': 'Cajas',
    'Total peso Bruto': 'Total peso Bruto (kg)',
    'Total peso Neto': 'Total peso Neto (kg)',
    'Total Volumen': 'Total volumen',
    'Total Importe': 'Ventas totales FOB',
    'Total LibrUtiliz': 'Total Cajas'
}

def exportar_a_excel(contenedores_volumenes, mensajes_contenedores, df_final, columnas_mapeadas):
    try:
        if os.name == 'nt':  # Para Windows
            desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        else:  # Para macOS y Linux
            desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')

        base_filename = 'contenedores_exportados'
        version = 1
        filename = f'{base_filename}({version}).xlsx'
        filepath = os.path.join(desktop_path, filename)

        while os.path.exists(filepath):
            version += 1
            filename = f'{base_filename}({version}).xlsx'
            filepath = os.path.join(desktop_path, filename)

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for i, (volumen_contenedor, mensajes) in enumerate(zip(contenedores_volumenes, mensajes_contenedores)):
                df_contenedor = df_final.loc[mensajes].copy()

                df_contenedor.rename(columns={
                    'Bruto': 'Peso bruto (kg)',
                    'Neto': 'Peso neto (kg)',
                    'Volumen': 'Volumen (m3)',
                    'Importe': 'Valor FOB',
                    'Doc.comer.': 'Pedido',
                    'LibrUtiliz': 'Cajas'
                }, inplace=True)

                # Asegurarse de que la columna 'Omitido' existe
                if 'Omitido' not in df_contenedor.columns:
                    df_contenedor['Omitido'] = ''

                # Reordenar las columnas
                columnas_ordenadas = ['Material', 'Texto de mensaje', 'Cajas', 'Peso bruto (kg)', 'Peso neto (kg)',
                                      'Volumen (m3)', 'Valor FOB', 'Cliente', 'Nombre', 'Contador', 'Pedido',
                                      'Grupo', 'Lote', 'Omitido']

                # Verificar que todas las columnas existen en df_contenedor
                columnas_existentes = [col for col in columnas_ordenadas if col in df_contenedor.columns]
                df_contenedor = df_contenedor[columnas_existentes]

                hoja_nombre = f'Contenedor_{i+1}'
                df_contenedor.to_excel(writer, sheet_name=hoja_nombre, index=False)

                workbook = writer.book
                worksheet = writer.sheets[hoja_nombre]

                worksheet.sheet_view.showGridLines = False

                title_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
                title_font = Font(color="FFFFFF", bold=True, size=14)
                title_alignment = Alignment(horizontal="center", vertical="center")

                for col in range(1, len(df_contenedor.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = title_fill
                    cell.font = title_font
                    cell.alignment = title_alignment

                # Si la columna 'Omitido' existe, configurar su encabezado y estilo
                if 'Omitido' in df_contenedor.columns:
                    col_idx = df_contenedor.columns.get_loc('Omitido') + 1  # Índice basado en 1 para openpyxl
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.value = 'Omitido'
                    cell.fill = title_fill
                    cell.font = title_font
                    cell.alignment = title_alignment

                total_label_col = 'O'  # Columna 15 (O)
                total_value_col = 'P'  # Columna 16 (P)

                bold_font_total = Font(bold=True, size=14)
                alignment_total = Alignment(horizontal="center", vertical="center")

                total_names = [
                    'Total peso Bruto (kg)',
                    'Total peso Neto (kg)',
                    'Total volumen',
                    'Ventas totales FOB',
                    'Total Cajas'
                ]
                totals_start_row = df_contenedor.shape[0] + 2

                for idx, total_name in enumerate(total_names):
                    cell_row = totals_start_row + idx
                    label_cell = worksheet[f'{total_label_col}{cell_row}']
                    value_cell = worksheet[f'{total_value_col}{cell_row}']

                    label_cell.value = total_name
                    label_cell.font = bold_font_total
                    label_cell.alignment = alignment_total

                    # Actualizar las fórmulas con el número correcto de filas
                    end_row = 100
                    if total_name == 'Total peso Bruto (kg)':
                        formula = f"=SUMPRODUCT((LOWER(${get_column_letter(df_contenedor.columns.get_loc('Omitido') + 1)}$2:${get_column_letter(df_contenedor.columns.get_loc('Omitido') + 1)}${end_row})<>\"x\")*(${get_column_letter(df_contenedor.columns.get_loc('Peso bruto (kg)') + 1)}2:${get_column_letter(df_contenedor.columns.get_loc('Peso bruto (kg)') + 1)}{end_row})*(${get_column_letter(df_contenedor.columns.get_loc('Cajas') + 1)}2:${get_column_letter(df_contenedor.columns.get_loc('Cajas') + 1)}{end_row}))"
                    elif total_name == 'Total peso Neto (kg)':
                        formula = f"=SUMPRODUCT((LOWER(${get_column_letter(df_contenedor.columns.get_loc('Omitido') + 1)}$2:${get_column_letter(df_contenedor.columns.get_loc('Omitido') + 1)}${end_row})<>\"x\")*(${get_column_letter(df_contenedor.columns.get_loc('Peso neto (kg)') + 1)}2:${get_column_letter(df_contenedor.columns.get_loc('Peso neto (kg)') + 1)}{end_row})*(${get_column_letter(df_contenedor.columns.get_loc('Cajas') + 1)}2:${get_column_letter(df_contenedor.columns.get_loc('Cajas') + 1)}{end_row}))"
                    elif total_name == 'Total volumen':
                        formula = f"=SUMPRODUCT((LOWER(${get_column_letter(df_contenedor.columns.get_loc('Omitido') + 1)}$2:${get_column_letter(df_contenedor.columns.get_loc('Omitido') + 1)}${end_row})<>\"x\")*(${get_column_letter(df_contenedor.columns.get_loc('Volumen (m3)') + 1)}2:${get_column_letter(df_contenedor.columns.get_loc('Volumen (m3)') + 1)}{end_row})*(${get_column_letter(df_contenedor.columns.get_loc('Cajas') + 1)}2:${get_column_letter(df_contenedor.columns.get_loc('Cajas') + 1)}{end_row}))"
                    elif total_name == 'Ventas totales FOB':
                        formula = f"=SUMPRODUCT((LOWER(${get_column_letter(df_contenedor.columns.get_loc('Omitido') + 1)}$2:${get_column_letter(df_contenedor.columns.get_loc('Omitido') + 1)}${end_row})<>\"x\")*(${get_column_letter(df_contenedor.columns.get_loc('Valor FOB') + 1)}2:${get_column_letter(df_contenedor.columns.get_loc('Valor FOB') + 1)}{end_row})*(${get_column_letter(df_contenedor.columns.get_loc('Contador') + 1)}2:${get_column_letter(df_contenedor.columns.get_loc('Contador') + 1)}{end_row})*(${get_column_letter(df_contenedor.columns.get_loc('Cajas') + 1)}2:${get_column_letter(df_contenedor.columns.get_loc('Cajas') + 1)}{end_row}))"
                    elif total_name == 'Total Cajas':
                        formula = f"=SUMPRODUCT((LOWER(${get_column_letter(df_contenedor.columns.get_loc('Omitido') + 1)}$2:${get_column_letter(df_contenedor.columns.get_loc('Omitido') + 1)}${end_row})<>\"x\")*(${get_column_letter(df_contenedor.columns.get_loc('Cajas') + 1)}2:${get_column_letter(df_contenedor.columns.get_loc('Cajas') + 1)}{end_row}))"

                    value_cell.value = formula
                    value_cell.font = bold_font_total
                    value_cell.alignment = alignment_total

                # Agregar la fila combinada "Adición De Referencias"
                merged_row = totals_start_row + len(total_names) + 2  # Dos filas debajo de los totales
                worksheet.merge_cells(start_row=merged_row, start_column=1, end_row=merged_row, end_column=13)  # Columnas A (1) a M (13)
                merged_cell = worksheet.cell(row=merged_row, column=1)
                merged_cell.value = 'Adición De Referencias'
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                merged_cell.font = Font(bold=True, size=14)

                # Establecer el color violeta claro
                violet_fill = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")
                for col in range(1, 14):  # Columnas A (1) a M (13)
                    cell = worksheet.cell(row=merged_row, column=col)
                    cell.fill = violet_fill

                # Ajustar el ancho de las columnas
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column].width = adjusted_width

        filas_exportadas = sum(len(mensajes) for mensajes in mensajes_contenedores)
        total_filas = df_final.shape[0]
        if filas_exportadas != total_filas:
            print(f"⚠️ Error de Exportación: Se han exportado {filas_exportadas} filas, pero el total es {total_filas}.")
        else:
            print(f"✅ Todas las filas han sido exportadas correctamente a Excel: {filename}")

    except Exception as e:
        print(f"Error durante la exportación: {str(e)}")

def calcular_totales(df):
    # Asegurar que las columnas sean numéricas
    for col in ['Bruto', 'Neto', 'Volumen', 'Importe', 'LibrUtiliz', 'Contador']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # Calcular totales multiplicando por 'LibrUtiliz' donde corresponda
    total_bruto = round((df['Bruto'] * df['LibrUtiliz']).sum(), 2)
    total_neto = round((df['Neto'] * df['LibrUtiliz']).sum(), 2)
    total_volumen = round((df['Volumen'] * df['LibrUtiliz']).sum(), 2)
    total_importe = round((df['Importe'] * df['Contador'] * df['LibrUtiliz']).sum(), 2)
    total_librUtiliz = round(df['LibrUtiliz'].sum(), 2)

    return {
        'Total peso Bruto': total_bruto,
        'Total peso Neto': total_neto,
        'Total Volumen': total_volumen,
        'Total Importe': total_importe,
        'Total LibrUtiliz': total_librUtiliz
    }

def calcular_contenedores(df, columnas_mapeadas, capacidad_contenedor_max):
    # Ordenar el DataFrame por la columna 'Lote' de menor a mayor (manteniendo como string)
    df = df.sort_values(by='Lote')

    # Inicializar listas de contenedores y mensajes
    contenedores_volumenes = []
    mensajes_contenedores = []

    # Crear una nueva columna 'Volumen_LibrUtiliz'
    df['Volumen_LibrUtiliz'] = df['Volumen'] * df['LibrUtiliz']
    df['Color_Volumen'] = ''

    # Inicializar variables
    current_container = []
    cumulative_vol = 0
    total_rows = df.shape[0]
    processed_rows = 0

    for index, row in df.iterrows():
        processed_rows += 1
        vol_lib_util = row['Volumen_LibrUtiliz']
        mensaje = index  # Usar el índice único como identificador

        # Si el volumen del mensaje excede la capacidad, creamos un contenedor para él solo
        if vol_lib_util > capacidad_contenedor_max:
            # Marcar la fila para colorear
            df.at[index, 'Color_Volumen'] = 'yellow'
            # Incluir la fila en su propio contenedor
            contenedores_volumenes.append(vol_lib_util)
            mensajes_contenedores.append([mensaje])
            print(f"Fila {processed_rows}/{total_rows} asignada a Contenedor individual por exceso de volumen.")
            continue

        # Si el mensaje no cabe en el contenedor actual, llenamos el contenedor y empezamos uno nuevo
        if cumulative_vol + vol_lib_util > capacidad_contenedor_max:
            # Finalizar contenedor actual
            if current_container:
                contenedores_volumenes.append(cumulative_vol)
                mensajes_contenedores.append(current_container)
                print(f"Contenedor finalizado con volumen total: {cumulative_vol:.2f}")

            # Iniciar un nuevo contenedor con la fila actual
            current_container = [mensaje]
            cumulative_vol = vol_lib_util
            print(f"Fila {processed_rows}/{total_rows} iniciando un nuevo Contenedor. Volumen: {cumulative_vol:.2f}")
        else:
            # Agregar a contenedor actual
            current_container.append(mensaje)
            cumulative_vol += vol_lib_util
            print(f"Fila {processed_rows}/{total_rows} agregada al Contenedor actual. Volumen acumulado: {cumulative_vol:.2f}")

    # Si hay un contenedor actual al final, agregarlo
    if current_container:
        contenedores_volumenes.append(cumulative_vol)
        mensajes_contenedores.append(current_container)
        print(f"Último Contenedor finalizado con volumen total: {cumulative_vol:.2f}")

    print(f"Total de filas procesadas: {processed_rows}/{total_rows}")
    print(f"Total de contenedores creados: {len(contenedores_volumenes)}")

    # Validar que todas las filas hayan sido asignadas a algún contenedor
    filas_asignadas = sum(len(mensajes) for mensajes in mensajes_contenedores)
    total_filas = df.shape[0]
    if filas_asignadas != total_filas:
        print(f"⚠️ Error: Se han asignado {filas_asignadas} filas, pero el total es {total_filas}.")
    else:
        print("✅ Todas las filas han sido asignadas correctamente a contenedores.")

    return contenedores_volumenes, mensajes_contenedores

def mostrar_resultados(totales, contenedores_volumenes, mensajes_contenedores, df_final, columnas_mapeadas, capacidad_contenedor_max):
    root = tk.Tk()
    root.title("Resultados de Contenedores")
    root.geometry("900x600")  # Ajuste del tamaño de la ventana

    # Configurar estilo
    style = ttk.Style(root)
    style.theme_use("clam")  # Puedes cambiar el tema según preferencia

    # Definir estilos personalizados
    style.configure("Treeview.Heading", font=("Arial", 14, "bold"), background="#0000FF", foreground="white")
    style.configure("Treeview", font=("Arial", 14), rowheight=30, fieldbackground="#f0f0f0")
    style.map("Treeview", background=[("selected", "#ADD8E6")], foreground=[("selected", "black")])

    # Frame para Totales
    frame_totales = ttk.LabelFrame(root, text="Totales", padding=(20, 10))
    frame_totales.pack(fill='x', padx=20, pady=10)

    # Treeview para Totales
    columnas_totales = ("Descripción", "Valor")
    tree_totales = ttk.Treeview(frame_totales, columns=columnas_totales, show="headings", height=5)
    tree_totales.heading("Descripción", text="Descripción")
    tree_totales.heading("Valor", text="Valor")
    tree_totales.column("Descripción", anchor='w', width=300)
    tree_totales.column("Valor", anchor='center', width=300)

    # Insertar datos de totales
    for desc, valor in totales.items():
        # Usar el mapeo para los nombres de totales
        desc_mapeado = COLUMN_MAPPING.get(desc, desc)
        tree_totales.insert("", "end", values=(desc_mapeado, f"{valor:,.2f}"))

    # Scrollbar para Totales
    scrollbar_totales = ttk.Scrollbar(frame_totales, orient="vertical", command=tree_totales.yview)
    tree_totales.configure(yscroll=scrollbar_totales.set)
    scrollbar_totales.pack(side='right', fill='y')
    tree_totales.pack(fill='x')

    # Frame para Contenedores
    frame_contenedores = ttk.LabelFrame(root, text="Contenedores", padding=(20, 10))
    frame_contenedores.pack(fill='both', expand=True, padx=20, pady=10)

    # Treeview para Contenedores
    columnas_contenedores = ("Número de Contenedor", "Peso Neto (unidades)")
    tree_contenedores = ttk.Treeview(frame_contenedores, columns=columnas_contenedores, show="headings", height=5)
    tree_contenedores.heading("Número de Contenedor", text="Número de Contenedor")
    tree_contenedores.heading("Peso Neto (unidades)", text="Peso Neto (unidades)")

    tree_contenedores.column("Número de Contenedor", anchor='center', width=300)
    tree_contenedores.column("Peso Neto (unidades)", anchor='center', width=300)

    # Insertar datos de contenedores
    for i, contenedor in enumerate(contenedores_volumenes, start=1):
        tree_contenedores.insert("", "end", iid=i-1, values=(f"Contenedor {i}", f"{contenedor:,.2f}"))

    # Scrollbar para Contenedores
    scrollbar_contenedores = ttk.Scrollbar(frame_contenedores, orient="vertical", command=tree_contenedores.yview)
    tree_contenedores.configure(yscroll=scrollbar_contenedores.set)
    scrollbar_contenedores.pack(side='right', fill='y')
    tree_contenedores.pack(fill='both', expand=True)

    def on_select(event):
        selected_item = tree_contenedores.selection()
        if not selected_item:
            return

        contenedor_index = int(selected_item[0])

        # Crear una nueva ventana para mostrar los mensajes
        mensajes_ventana = tk.Toplevel(root)
        mensajes_ventana.title(f"Mensajes para Contenedor {contenedor_index + 1}")
        mensajes_ventana.geometry("800x400")  # Ajuste del tamaño de la ventana de mensajes

        # Configurar estilo para la nueva ventana
        style_mensajes = ttk.Style(mensajes_ventana)
        style_mensajes.theme_use("clam")
        style_mensajes.configure("Treeview.Heading", font=("Arial", 14, "bold"), background="#0000FF", foreground="white")
        style_mensajes.configure("Treeview", font=("Arial", 14), rowheight=30, fieldbackground="#f0f0f0")
        style_mensajes.map("Treeview", background=[("selected", "#ADD8E6")], foreground=[("selected", "black")])

        # Frame para los mensajes
        frame_mensajes = ttk.Frame(mensajes_ventana, padding=(20, 10))
        frame_mensajes.pack(fill='both', expand=True)

        # Definir las columnas que deseas mostrar, incluyendo 'Lote'
        columnas_mensajes = list(df_final.columns)
        if 'Lote' not in columnas_mensajes:
            columnas_mensajes.append('Lote')  # Añadir 'Lote' si no está presente

        # Treeview para Mensajes con todas las columnas
        tree_mensajes = ttk.Treeview(frame_mensajes, columns=columnas_mensajes, show="headings", height=15)

        # Definir encabezados y columnas
        for col in columnas_mensajes:
            # Usar el mapeo para los encabezados
            encabezado = COLUMN_MAPPING.get(col, col)
            tree_mensajes.heading(col, text=encabezado)
            tree_mensajes.column(col, anchor='center', width=150)

        # Insertar datos de mensajes
        for mensaje_idx in mensajes_contenedores[contenedor_index]:
            mensaje = df_final.loc[mensaje_idx]
            valores = [mensaje[col] for col in columnas_mensajes]
            tree_mensajes.insert("", "end", values=valores)

        # Scrollbars para Mensajes
        scrollbar_mensajes_y = ttk.Scrollbar(frame_mensajes, orient="vertical", command=tree_mensajes.yview)
        scrollbar_mensajes_x = ttk.Scrollbar(frame_mensajes, orient="horizontal", command=tree_mensajes.xview)
        tree_mensajes.configure(yscroll=scrollbar_mensajes_y.set, xscroll=scrollbar_mensajes_x.set)
        scrollbar_mensajes_y.pack(side='right', fill='y')
        scrollbar_mensajes_x.pack(side='bottom', fill='x')
        tree_mensajes.pack(side='left', fill='both', expand=True)

    # Bind del evento de selección
    tree_contenedores.bind("<<TreeviewSelect>>", on_select)

    # Botón para Exportar a Excel
    boton_exportar = ttk.Button(root, text="Exportar a Excel", command=lambda: exportar_a_excel(
        contenedores_volumenes, mensajes_contenedores, df_final, columnas_mapeadas
    ))

    boton_exportar.pack(pady=10)

    # Verificación de integridad antes de iniciar el loop
    filas_asignadas = sum(len(mensajes) for mensajes in mensajes_contenedores)
    total_filas = df_final.shape[0]
    if filas_asignadas != total_filas:
        messagebox.showerror("Error de Integridad", f"Se han asignado {filas_asignadas} filas, pero el total es {total_filas}.")
        print(f"⚠️ Error: Se han asignado {filas_asignadas} filas, pero el total es {total_filas}.")
    else:
        print("✅ Todas las filas han sido asignadas correctamente a contenedores.")

    root.mainloop()

def main_proceso(df_final, columnas_mapeadas, capacidad_contenedor_max):
    try:
        # Verificar que las columnas requeridas estén presentes
        required_columns = ['Material', 'Texto de mensaje', 'Bruto', 'Neto', 'Volumen', 'Importe',
                            'LibrUtiliz', 'Contador', 'Cliente', 'Nombre', 'Doc.comer.', 'Grupo', 'Lote']
        for col in required_columns:
            if col not in df_final.columns:
                messagebox.showerror("Error", f"La columna '{col}' no está presente en los datos.")
                print(f"❌ Error: La columna '{col}' no está presente en los datos.")
                return

        # Rellenar valores nulos en la columna 'Nombre' si es necesario
        df_final['Nombre'] = df_final['Nombre'].fillna('Sin Nombre')

        # Calcular contenedores
        contenedores_volumenes, mensajes_contenedores = calcular_contenedores(df_final, columnas_mapeadas, capacidad_contenedor_max)

        # Calcular totales
        totales = calcular_totales(df_final)

        # Mostrar resultados
        mostrar_resultados(
            totales, contenedores_volumenes, mensajes_contenedores, df_final, columnas_mapeadas, capacidad_contenedor_max
        )
    except Exception as e:
        messagebox.showerror("Error", f"Error en el proceso principal: {e}")
        print(f"❌ Error en el proceso principal: {e}")

def cargar_datos(ruta_archivo):
    try:
        df = pd.read_excel(ruta_archivo)
        print(f"✅ Datos cargados exitosamente desde '{ruta_archivo}'.")
        return df
    except Exception as e:
        messagebox.showerror("Error de Carga", f"Ocurrió un error al cargar el archivo: {e}")
        print(f"❌ Error al cargar el archivo: {e}")
        return None

# Si deseas mantener una función main en operaciones.py, asegúrate de que esté completa.
def main():
    try:
        # Crear la ventana principal para seleccionar el archivo de datos
        root = tk.Tk()
        root.withdraw()  # Ocultar la ventana principal

        messagebox.showinfo("Seleccionar Archivo", "Seleccione el archivo Excel que contiene los datos.")

        ruta_archivo = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos de Excel", "*.xlsx *.xls")]
        )

        if not ruta_archivo:
            messagebox.showwarning("Sin Selección", "No se ha seleccionado ningún archivo. El programa se cerrará.")
            print("⚠️ No se ha seleccionado ningún archivo. El programa se cerrará.")
            return

        df_final = cargar_datos(ruta_archivo)
        if df_final is None:
            return

        # Definir las columnas mapeadas si es necesario (ejemplo)
        columnas_mapeadas = {
            'Bruto': 'Peso bruto (kg)',
            'Neto': 'Peso neto (kg)',
            'Volumen': 'Volumen (m3)',
            'Importe': 'Valor FOB',
            'Doc.comer.': 'Pedido',
            'LibrUtiliz': 'Cajas',
            'Total peso Bruto': 'Total peso Bruto (kg)',
            'Total peso Neto': 'Total peso Neto (kg)',
            'Total Volumen': 'Total volumen',
            'Total Importe': 'Ventas totales FOB',
            'Total LibrUtiliz': 'Total Cajas'
        }

        # Definir la capacidad máxima del contenedor (ejemplo)
        capacidad_contenedor_max = 1000  # Ajusta este valor según tus necesidades

        # Llamar a la función principal de procesamiento
        main_proceso(df_final, columnas_mapeadas, capacidad_contenedor_max)
    except Exception as e:
        messagebox.showerror("Error", f"Error en la función main: {e}")
        print(f"❌ Error en la función main: {e}")

if __name__ == "__main__":
    main()
